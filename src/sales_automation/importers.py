from __future__ import annotations

import csv
import base64
import hashlib
import io
from typing import Any


FIELD_ALIASES = {
    "linkedin_url": ["linkedin_url", "linkedin", "linkedin profile", "linkedin profile url", "profile_url", "url"],
    "first_name": ["first_name", "first name", "firstname", "given_name"],
    "last_name": ["last_name", "last name", "lastname", "family_name"],
    "email": ["email", "work_email", "work email", "business_email"],
    "phone": ["phone", "mobile", "telephone", "tel", "phone_number", "phone number"],
    "job_title": ["job_title", "job title", "title", "role", "position"],
    "company_name": ["company_name", "company", "company name", "organization", "account"],
    "company_domain": ["company_domain", "domain", "company domain", "website", "company_website"],
    "industry": ["industry", "sector"],
    "location": ["location", "city", "country", "region"],
    "notes": ["notes", "note"],
    "source": ["source"],
}

COMPANY_SEED_ALIASES = {
    "company_name": ["company_name", "company", "company name", "store_name", "store name", "公司", "公司名称", "公司/店铺名称", "店铺名称"],
    "category": ["category", "类别", "类目", "行业类别", "业务类别"],
    "reason": ["reason", "background", "research", "notes", "简短背调", "公司背调", "实力和匹配度", "Vertu渠道价值", "简短背调(为何匹配Vertu资质和清理 & 调性契合度)", "简短背调_为何匹配vertu资质和清理_调性契合度"],
    "website": ["website", "domain", "company_domain", "company website", "官网", "官网/联系链接", "官网链接", "联系链接", "公司网址"],
    "job_titles": ["job_titles", "job titles", "titles", "roles", "role", "position", "职位", "目标职位"],
    "industry": ["industry", "sector", "行业"],
    "location": ["location", "country", "region", "地区", "国家", "主要城市", "城市", "公司地址"],
    "phone": ["phone", "telephone", "mobile", "电话", "手机号", "联系电话"],
    "email": ["email", "邮箱", "联系邮箱"],
}


def parse_contacts_csv(text: str, *, default_source: str = "csv_import") -> list[dict[str, Any]]:
    stream = io.StringIO(text)
    sample = text[:2048]
    dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
    reader = csv.DictReader(stream, dialect=dialect)
    if not reader.fieldnames:
        return []
    mapping = _build_mapping(reader.fieldnames)
    contacts: list[dict[str, Any]] = []
    for row in reader:
        contact: dict[str, Any] = {}
        for target, source in mapping.items():
            value = (row.get(source) or "").strip()
            if value:
                contact[target] = value
        if not contact.get("linkedin_url"):
            contact["linkedin_url"] = _synthetic_url(contact)
        if contact.get("company_domain"):
            contact["company_domain"] = _normalize_domain(contact["company_domain"])
        if contact.get("email"):
            contact["email_status"] = "valid"
            contact.setdefault("status", "enriched")
        if contact.get("phone"):
            contact["phone_candidates"] = [{"phone": contact["phone"], "source": default_source, "status": "provided"}]
        contact.setdefault("source", default_source)
        if _has_minimum_identity(contact):
            contacts.append(contact)
    return contacts


def parse_company_seed_csv(text: str, *, default_location: str = "", default_industry: str = "") -> list[dict[str, Any]]:
    stream = io.StringIO(text)
    sample = text[:2048]
    dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
    reader = csv.DictReader(stream, dialect=dialect)
    if not reader.fieldnames:
        return []
    mapping = _build_mapping_for_aliases(reader.fieldnames, COMPANY_SEED_ALIASES)
    seeds: list[dict[str, Any]] = []
    for row in reader:
        seed: dict[str, Any] = {}
        for target, source in mapping.items():
            value = (row.get(source) or "").strip()
            if value:
                seed[target] = value
        if seed.get("website"):
            seed["company_domain"] = _normalize_domain(seed["website"])
        elif seed.get("company_domain"):
            seed["company_domain"] = _normalize_domain(seed["company_domain"])
        if seed.get("job_titles"):
            seed["job_titles"] = _split_job_titles(seed["job_titles"])
        else:
            seed["job_titles"] = []
        seed.setdefault("industry", default_industry or seed.get("category") or "")
        seed.setdefault("location", default_location)
        if seed.get("phone"):
            seed["phone_candidates"] = [{"phone": seed["phone"], "source": "company_seed", "status": "provided"}]
        if seed.get("email"):
            seed["email_candidates"] = [{"email": seed["email"], "source": "company_seed", "status": "provided", "category": "company_generic", "confidence": 50}]
        if seed.get("company_name") or seed.get("company_domain"):
            seeds.append(seed)
    return seeds


def parse_company_seed_upload(
    *,
    filename: str,
    content_base64: str,
    default_location: str = "",
    default_industry: str = "",
) -> list[dict[str, Any]]:
    content = base64.b64decode(content_base64)
    lowered = filename.lower()
    if lowered.endswith(".csv"):
        text = content.decode("utf-8-sig")
        return parse_company_seed_csv(text, default_location=default_location, default_industry=default_industry)
    if lowered.endswith(".xlsx") or lowered.endswith(".xlsm"):
        rows = _read_xlsx_rows(content)
        return _parse_company_seed_rows(rows, default_location=default_location, default_industry=default_industry)
    raise ValueError("unsupported_import_file_type")


def _read_xlsx_rows(content: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl_missing") from exc
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    raw_rows = [[_cell_to_text(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]
    header_index = _find_header_row(raw_rows, COMPANY_SEED_ALIASES)
    if header_index < 0:
        return []
    headers = raw_rows[header_index]
    rows: list[dict[str, str]] = []
    for values in raw_rows[header_index + 1 :]:
        row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers)) if headers[index]}
        if any(str(value or "").strip() for value in row.values()):
            rows.append(row)
    return rows


def _parse_company_seed_rows(rows: list[dict[str, str]], *, default_location: str = "", default_industry: str = "") -> list[dict[str, Any]]:
    if not rows:
        return []
    fieldnames = list(rows[0].keys())
    mapping = _build_mapping_for_aliases(fieldnames, COMPANY_SEED_ALIASES)
    seeds: list[dict[str, Any]] = []
    for row in rows:
        seed: dict[str, Any] = {}
        for target, source in mapping.items():
            value = (row.get(source) or "").strip()
            if value:
                seed[target] = value
        if seed.get("website"):
            seed["company_domain"] = _normalize_domain(seed["website"])
        elif seed.get("company_domain"):
            seed["company_domain"] = _normalize_domain(seed["company_domain"])
        if seed.get("job_titles"):
            seed["job_titles"] = _split_job_titles(seed["job_titles"])
        else:
            seed["job_titles"] = []
        seed.setdefault("industry", default_industry or seed.get("category") or "")
        seed.setdefault("location", default_location)
        if seed.get("phone"):
            seed["phone_candidates"] = [{"phone": seed["phone"], "source": "company_seed", "status": "provided"}]
        if seed.get("email"):
            seed["email_candidates"] = [{"email": seed["email"], "source": "company_seed", "status": "provided", "category": "company_generic", "confidence": 50}]
        if seed.get("company_name") or seed.get("company_domain"):
            seeds.append(seed)
    return seeds


def _build_mapping(fieldnames: list[str]) -> dict[str, str]:
    return _build_mapping_for_aliases(fieldnames, FIELD_ALIASES)


def _build_mapping_for_aliases(fieldnames: list[str], aliases_by_target: dict[str, list[str]]) -> dict[str, str]:
    normalized = {_normalize(name): name for name in fieldnames}
    mapping: dict[str, str] = {}
    for target, aliases in aliases_by_target.items():
        for alias in aliases:
            if _normalize(alias) in normalized:
                mapping[target] = normalized[_normalize(alias)]
                break
    return mapping


def _normalize(value: str) -> str:
    return re_sub_non_word(value.strip().lower())


def _find_header_row(rows: list[list[str]], aliases_by_target: dict[str, list[str]]) -> int:
    best_index = -1
    best_score = 0
    for index, row in enumerate(rows[:20]):
        fieldnames = [value for value in row if value]
        if not fieldnames:
            continue
        mapping = _build_mapping_for_aliases(fieldnames, aliases_by_target)
        score = len(mapping)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index if best_score >= 1 else -1


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_domain(value: str) -> str:
    value = value.strip()
    value = value.removeprefix("https://").removeprefix("http://").split("/")[0]
    if value.startswith("www."):
        value = value[4:]
    return value


def _split_job_titles(value: str) -> list[str]:
    return [item.strip(" '\"\t\r\n") for item in value.replace("，", ",").replace("；", ",").replace(";", ",").split(",") if item.strip(" '\"\t\r\n")]


def re_sub_non_word(value: str) -> str:
    return value.replace("-", "_").replace(" ", "_").replace("/", "_").replace("&", "_").replace("(", "").replace(")", "")


def _synthetic_url(contact: dict[str, Any]) -> str:
    seed = "|".join(
        [
            contact.get("email", ""),
            contact.get("first_name", ""),
            contact.get("last_name", ""),
            contact.get("company_name", ""),
            contact.get("company_domain", ""),
        ]
    )
    digest = hashlib.sha256(seed.encode("utf-8")).hexdigest()[:24]
    return f"manual://csv/{digest}"


def _has_minimum_identity(contact: dict[str, Any]) -> bool:
    return bool(contact.get("email") or contact.get("company_name") or contact.get("company_domain") or contact.get("linkedin_url"))
